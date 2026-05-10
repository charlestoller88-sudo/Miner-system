import argparse
import csv
import json
import os
import re
import sys
import urllib.error
import urllib.request
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

try:
    from openpyxl import Workbook
except ImportError:
    Workbook = None


# 项目根目录：fault_patterns_learned.json 固定放在此目录（与独立版「AI诊断矿机运行日志报告」一致）
SCRIPT_DIR = Path(__file__).resolve().parent.parent
THRESHOLD_THS = 50.0
# 0 算力浮点容差：避免 0.00000x 的解析噪音
ZERO_THS_EPSILON = float(os.environ.get("MINER_ZERO_THS_EPSILON", "0.05"))
# 近零算力上限：命中则归类为「无算力（近零）」而非「低算力」
NEAR_ZERO_THS_ABS = float(os.environ.get("MINER_NEAR_ZERO_THS_ABS", "2.0"))
# 额定算力比例阈值：近零阈值 = max(绝对阈值, NameplateTHS * 比例)
NEAR_ZERO_NAMEPLATE_RATIO = float(os.environ.get("MINER_NEAR_ZERO_NAMEPLATE_RATIO", "0.03"))
OLLAMA_URL = "http://127.0.0.1:11434/api/generate"
OLLAMA_MODEL = os.environ.get("MINER_AI_MODEL", "qwen3:30b")
MAX_AI_CALLS = int(os.environ.get("MINER_AI_MAX_CALLS", "8"))
# 设为 1 时：即使规则引擎已给出主因，仍用本地 LLM 覆盖（默认 0：规则优先）
FORCE_LLM_OVERRIDE = os.environ.get("MINER_AI_FORCE_LLM", "").strip() in ("1", "true", "True")
# 矿池/Stratum 频繁断连：三类典型报错合计命中次数 ≥ 此值则优先于「风扇」等高频词
STRATUM_UNSTABLE_MIN_HITS = int(os.environ.get("MINER_STRATUM_UNSTABLE_MIN", "8"))
# CONFIG 断电时：pool inactivity 占 Stratum 三类报错的比例 ≥ 此值则不把「矿池/Stratum」列为主因（多为停挖/空闲踢线）
STRATUM_POOL_INACTIVITY_DOMINANCE = float(os.environ.get("MINER_POOL_INACTIVITY_RATIO", "0.45"))
# LuxOS「FANx: ... -> FAIL」至少出现次数才视为风扇反复故障（与 IsPowerSupplyOn:false 联判）
LUXOS_FAN_FAIL_MIN_TRANSITIONS = int(os.environ.get("MINER_LUXOS_FAN_FAIL_MIN", "3"))

# LuxOS/Stratum：连接质量差（与 fan 词频解耦，需单独计分）
# 规则引擎「明确优先级路径」：用于 confidence=高（learned 单独为 中）
PRIORITY_HIGH_LANES = frozenset(
    {
        "curtail",
        "tsensor",
        "fan",
        "fan_luxos",
        "power_cfg",
        "tuner",
        "soc",
        "hashboard",
        "init",
        "pll",
        "stratum",
        "eeprom_e4",
        "hashchip",
        "temp_diff",
        "eeprom_init",
    }
)

STRATUM_UNSTABLE_PATTERNS = [
    r"connection dropped due to pool inactivity",
    r"stratum read failure",
    r"connection closed by the server",
]

RULES = [
    # 勿单独匹配 asic：DEVS/SUMMARY 中大量出现易压过真正结构化主因
    (r"no active chains|hashboard .* problem|tuner error|unexpected revision|detected \d+\s+chips", "算力板/芯片异常",
     ["断电后重插算力板排线", "单板逐个测试定位坏板", "检查板卡供电与固件版本"]),
    # 勿用裸子串 dns：会匹配 CONFIG 中 "DNS Servers" 字段名，造成假阳性（真 DNS 故障用 failed to resolve / dns error 等）
    (
        r"failed to resolve|no active pools|stratum.*disconnected|"
        r"socket (?:error|timeout|closed|reset)|"
        r"dns (?:error|failure|timeout|lookup failed)|unable to resolve.*dns",
        "网络/矿池连接异常",
        ["检查 DNS 和网关配置", "切换备用矿池地址与端口", "检查交换机端口及网线"],
    ),
    (r"connection refused|deadline|api error|client disconnected", "矿机服务异常/接口超时",
     ["重启矿机服务或整机重启", "检查控制板负载和系统日志", "确认管理端口未被占用"]),
    # 勿使用裸关键词 power：会匹配「power off hashboard」等关机动作，误判为 PSU 故障
    (
        r"failed to detect psu|failed to set up psu|psu overcurrent|undervolt|over voltage|"
        r"low voltage|power supply.*(fail|error|fault|abnormal)",
        "电源异常",
        ["检查 PSU 供电和电压稳定性", "更换电源或供电回路", "排查电源线接触不良"],
    ),
    (
        r"error_fan_lost|fan lost,\s*only find|stop_mining:\s*fan|tachometer.*(error|fail)",
        "风扇异常",
        ["检查风扇转速与供电", "清灰并确认风道", "更换故障风扇"],
    ),
    (
        r"overtemp|temperature too high|temp alarm|temp diff too high|error_temp_too_low|error_temp_too_high",
        "温度异常",
        ["检查机房进风温度", "清理散热片和灰尘", "降低频率后观察恢复情况"],
    ),
    # 勿用裸 i2c：内核「i2c /dev entries」等会刷屏；勿用裸 crc：「EEPROM error: CRC 1ST REGION」会误命中
    (r"(?<!:)i2c(?:\s+|\s*:\s*).{0,120}?(?:\berror\b|\bfail\b|timeout|nack|nak)|\buat\b|reg crc error|\bspi\b",
     "总线/通信异常",
     ["检查控制板与算力板通信线", "检查是否有干扰或松动", "升级固件并复测通讯稳定性"]),
]

PRIMARY_CAUSE_SOP = {
    "矿池/Stratum连接不稳定（频繁断连）": [
        "检查矿机到矿池域名的网络：延迟、丢包、DNS（可 ping / traceroute / 换 DNS 对比）",
        "检查交换机端口、网线、上联链路；排除同网段广播风暴或限速",
        "在 LuxOS 中切换备用矿池或矿池线路，排除矿池侧短暂故障",
        "重启矿机或重载 Stratum 连接后观察 Accepted 与算力是否恢复",
        "核对矿池子账户与矿工名；若仍频繁断连可尝试升级/回退 Luxminer 固件对比",
    ],
    "部分算力板失效（初始化/ASIC枚举失败）": [
        "断电后检查故障算力板与控制板之间的电源线、数据线是否插紧、有无氧化",
        "将故障板与正常板互换槽位或交叉到另一台同型号矿机，判断是板卡还是接口/控制板问题",
        "在 LuxOS 中对异常板尝试手动复位或重新上电（若界面支持）",
        "若多次 enumeration 后仍被永久关闭，按硬件损坏流程送修或更换算力板",
    ],
    "单链算力板故障致 SOC 停机（芯片数异常/CRC/通信）": [
        "根据日志定位异常 Chain（如 find 0 asic、仅 find 部分数量、reg crc error）",
        "断电后重插该链算力板排线与供电，做与正常链互换槽位测试以区分板卡与接口",
        "若故障随板迁移则更换算力板；若固定在槽位则检查控制板与该槽位连线",
        "若固件支持可尝试暂时禁用故障链仅跑健康链（注意功耗与收益）",
    ],
    "多块算力板故障致 SOC 停机（芯片数严重不足/CRC）": [
        "日志中若多条 Chain 出现「only find … asic」远低于 110，且伴随 ERROR_SOC_INIT，多为多块板同时异常",
        "分别检查各故障链的排线、供电接口；与正常链交叉换槽位判断是板还是槽位/控制板",
        "确认 PSU 各组输出是否稳定，但勿将「power off hashboard」误判为 PSU 本体损坏",
        "必要时逐板更换或送修；若固件支持可尝试仅启用健康链临时恢复部分算力",
    ],
    "限电/策略休眠（Curtailment Sleep）": [
        "在 LuxOS/矿场管理端或 API 来源处确认是否下发了 curtail/sleep 或限电策略",
        "若需恢复挖矿：将 CurtailMode 改为非 Sleep，或取消对应限电/休眠任务",
        "退出休眠后等待矿池重连，观察 SUMMARY/DEVS 是否从 Dead 恢复为正常",
        "若恢复后仍无算力或矿池反复断连，再单独排查网络与矿池配置",
    ],
    "温度传感器通信失败（误报超温/温差保护停机）": [
        "根据日志确认 fail to read tsensor / pic temp 的 chain 号，检查该链算力板排线、PIC 传感器侧连接是否松动或氧化",
        "与正常链交叉换槽位：故障随板走则更换算力板；固定在槽位则查控制板接口与线缆",
        "若日志中 PCB/芯片温度读数明显低于上限仍报 ERROR_TEMP，多为传感器数据异常而非真实过热",
        "确认对应链供电稳定；固件层面勿轻易关闭温度保护，优先硬件修复或更换故障板",
    ],
    "算力板严重异常（芯片识别/通信错误，NoPIC 可能无法单禁板）": [
        "关注 Tuner 报错：如 detected … chips 远超 110、unexpected revision、ERR:I1/I2，多为该板通信或硬件损坏",
        "断电后重插故障 HB 对应算力板排线与供电；与正常板互换槽位判断是板还是接口",
        "NoPIC 机型可能无法单独禁用故障板，整机可能无法继续挖矿，需先修复或更换该板",
        "若混有历史 Failed to resolve，以导出时间附近 tuner/hashboard 报错为准；矿池 Alive 不等于当前无板级故障",
    ],
    "算力板全板失效（芯片未检出/初始化失败）": [
        "断电后重新插拔控制板到三块算力板的数据排线与电源接口",
        "重点检查 HB1/HB2/HB3 是否都出现 no chips detected 或 NoChipsDetected",
        "交叉测试控制板与算力板，确认是控制板链路问题还是算力板损坏",
        "检查 PSU 对算力板供电电压与负载稳定性（是否触发过流保护）",
        "若仍失败，优先更换故障算力板或控制板",
    ],
    "算力板部分失效（芯片数量严重不足）": [
        "定位异常算力板并单板测试",
        "检查该板排线、供电及散热状态",
        "复位或升级固件后复测芯片识别数量",
        "若芯片长期缺失，安排维修或更换算力板",
    ],
    "算力板部分失效（多块链 Dead，仅部分链工作）": [
        "根据末段 [DEVS] 确认 Dead / Alive 链编号；断电后检查 Dead 链对应算力板数据排线与电源接头",
        "将故障链与正常链互换槽位或交叉到同型号矿机，判断故障随板还是随槽位/控制板",
        "仅单链或部分链工作时总算力约为额定三分之一属常见现象；矿池 Alive 不能排除板级失效",
        "若链长期保持 Dead，更换算力板或送修；必要时导出含启动阶段的完整日志查初始化错误",
    ],
    "算力板 EEPROM 解析失败（ERR:E4，无法启动挖矿）": [
        "根据日志定位无法解码的算力板编号（如 hashboard #2）；断电后重插该板数据排线与电源线，换槽位交叉测试判断是否随板损坏",
        "EEPROM 数据异常（如全 0）多为板载存储或硬件故障，重复失败需更换算力板或送修",
        "若同时出现 Failed to set up PSU / dummy backend，先在板卡与排线确认后再单独测 PSU 协议与供电稳定性",
        "paused by dead pools / missing license 多为未成功启动挖矿的连带现象，优先解决 ERR:E4 与硬件后再看矿池与授权",
    ],
    "算力板 EEPROM 读取失败（多链初始化失败，SOC 无法启动）": [
        "根据 load chain … eeprom / Data load fail / eeprom load ret:-1 定位失败链；逐链重插数据排线与电源，换槽位交叉测试区分板卡与控制板",
        "EEPROM error: CRC … REGION、Fixture data load failed 多为数据损坏或通信问题；多链同时失败优先查共因（排线、控制板 EEPROM 通路、固件版本）",
        "出现 ERROR_SOC_INIT: basic init failed / stop_mining: basic init failed 时，勿先当总线泛化故障；可尝试重刷官方或兼容固件后再测",
        "若单链异常可换板；多链皆失败且排除连接后重点考虑控制板或批量硬件问题",
    ],
    "算力板芯片枚举不足（Hashchip 与 110 不符，已被禁用）": [
        "根据 Tuner error / CHAIN 日志确认 {HB:n} 与「number of responses … doesn't match chip count 110」对应链",
        "断电后重插故障板数据排线与电源接口；与正常链互换槽位判断是板卡还是槽位/控制板",
        "仅一链正常时总算力约为额定三分之一属常见现象；PSU I²C checksum 异常多为次要，板卡稳定后再复测电源",
        "若某链长期 0 响应或响应数远低于 110，按算力板硬件损坏或通信链路故障处理，更换板卡或送修",
    ],
    "电源异常": [
        "检查 PSU 输出电压、功率余量和接线",
        "确认是否存在过流保护触发或掉压",
        "更换电源或供电回路后复测",
    ],
    "网络/矿池连接异常": [
        "检查 DNS、网关、交换机端口和网线",
        "核对矿池地址/端口并切换备用矿池",
        "观察重连频率与拒绝率是否恢复正常",
    ],
    "风扇异常": [
        "检查风扇供电、转速和风道堵塞",
        "清灰后复测温度与算力",
        "更换异常风扇",
    ],
    "风扇异常（丢失/数量不足触发保护停机）": [
        "根据日志确认要求风扇数量与实际检测数量（如 only find 3 < 4）",
        "断电后检查四个风扇插头是否插紧、线序是否与控制板丝印一致，必要时对调插头判断是风扇还是接口",
        "确认风扇转速是否异常或为 0；更换不转或信号异常的风扇",
        "若硬件正常仍误报，检查控制板风扇检测电路；ERROR_SOC_INIT 若紧跟 fan lost，多为保护停机连带而非算力板主因",
    ],
    "风扇异常（LuxOS Fan status 反复 FAIL，伴电源关闭）": [
        "根据 Fan status changes 定位反复 FAIL 的风扇位号（如 FAN3），检查插头、线序与风扇本体；与正常位对调以区分风扇与控制板接口",
        "CONFIG 中 IsPowerSupplyOn 为 false 时先排除保护关机原因，修复风扇后再通过界面/API 尝试上电",
        "矿池 Alive 但大量 pool inactivity/断连多为未挖矿或空闲踢线，不宜优先当网络故障",
        "若仍频繁 FAIL，更换风扇或送修控制板风扇检测电路",
    ],
    "电源未开启或保护停机（CONFIG 断电）": [
        "在 LuxOS Web/API 确认 IsPowerSupplyOn、CurtailMode 与外部管理策略是否强制下电",
        "检查风扇、温度传感器与 PSU；保护解除后再观察 SUMMARY/DEVS 与矿池连接",
        "日志中 pool inactivity 类断连多为未持续提交份额所致，修复上电与挖矿后再判网络",
    ],
    "总线/通信异常": [
        "检查控制板与算力板通信线缆是否松动或损坏",
        "排查干扰与接触不良",
        "升级固件后复测链路稳定性",
    ],
    "算力板/芯片异常": [
        "断电后重插算力板排线",
        "单板逐个测试定位坏板",
        "检查板卡供电与固件版本",
    ],
    "矿机服务异常/接口超时": [
        "重启矿机服务或整机重启",
        "检查控制板负载和系统日志",
        "确认管理端口未被占用",
    ],
    "温度异常": [
        "检查机房进风温度",
        "清理散热片和灰尘",
        "降低频率后观察恢复情况",
    ],
    "算力板间温差过大（温度保护停机）": [
        "根据 STATS/日志确认温差异常链（chain）与 PCB/芯片温度分布；重点清理该链散热片与风道、检查硅脂与散热器接触",
        "与其它正常链互换槽位，判断是单链散热/板卡问题还是槽位风道问题；Fan check passed 时仍可能风道堵塞或单链过热",
        "等待 hashboard cool 后仍报 temp diff too high 时，勿先当总线故障；排除积灰、接触不良后再考虑降频或报修",
        "勿随意放宽固件温差阈值；持续异常需更换算力板或厂家检测",
    ],
    "未识别明确故障关键词": [
        "建议人工复核完整日志",
        "优先执行供电、线缆、算力板三项基础排查",
    ],
}


def _load_learned_json() -> dict:
    p = SCRIPT_DIR / "fault_patterns_learned.json"
    if not p.is_file():
        return {}
    try:
        return json.loads(p.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError):
        return {}


def _merged_primary_sop() -> Dict[str, List[str]]:
    base = {k: list(v) for k, v in PRIMARY_CAUSE_SOP.items()}
    learned = _load_learned_json()
    for k, v in (learned.get("sop_overrides") or {}).items():
        if isinstance(v, list) and v:
            base[k] = [str(x) for x in v]
    return base


def _collect_rule_hits(lower: str) -> List[Tuple[int, str, List[str]]]:
    hits: List[Tuple[int, str, List[str]]] = []
    for pattern, label, actions in RULES:
        c = len(re.findall(pattern, lower))
        if c:
            hits.append((c, label, actions))
    hits.sort(key=lambda x: x[0], reverse=True)
    return hits


def _build_alternate_causes(
    primary: str,
    hits: List[Tuple[int, str, List[str]]],
    secondary_list: List[str],
    max_items: int = 6,
) -> str:
    """候选根因：RULES 命中排序（排除主因）+ 短次因说明（去重）。"""
    seen = {primary}
    out: List[str] = []
    for _, lab, _ in hits:
        if lab in seen:
            continue
        seen.add(lab)
        out.append(lab)
        if len(out) >= max_items:
            return "；".join(out)
    for s in secondary_list:
        if not s or s in seen or s == primary:
            continue
        if len(s) > 72:
            continue
        seen.add(s)
        out.append(s)
        if len(out) >= max_items:
            break
    return "；".join(out)


def _diagnosis_confidence(
    primary: str,
    hits: List[Tuple[int, str, List[str]]],
    priority_lane: Optional[str],
    from_rules_generic: bool,
) -> str:
    if primary == "未识别明确故障关键词":
        return "低"
    if priority_lane == "learned":
        return "中"
    if priority_lane in PRIORITY_HIGH_LANES:
        return "高"
    if from_rules_generic:
        if len(hits) < 2:
            return "高"
        r0, r1 = hits[0][0], hits[1][0]
        ratio = r0 / max(1, r1)
        if ratio >= 2.0:
            return "高"
        if ratio >= 1.35:
            return "中"
        return "低"
    return "中"


def _finalize_diagnosis_result(
    primary: str,
    secondary: List[str],
    solutions: str,
    priority_lane: Optional[str],
    from_rules_generic: bool,
    hits: List[Tuple[int, str, List[str]]],
) -> Dict[str, str]:
    conf = _diagnosis_confidence(primary, hits, priority_lane, from_rules_generic)
    alt = _build_alternate_causes(primary, hits, secondary)
    return {
        "primary_cause": primary,
        "secondary_causes": "；".join(secondary),
        "solutions": solutions,
        "confidence": conf,
        "alternate_causes": alt,
    }


def _stratum_unstable_total_hits(lower: str) -> int:
    n = 0
    for pat in STRATUM_UNSTABLE_PATTERNS:
        n += len(re.findall(pat, lower))
    return n


def _stratum_pool_inactivity_hits(lower: str) -> int:
    return len(re.findall(r"connection dropped due to pool inactivity", lower))


def _luxos_fan_fail_transitions(lower: str) -> int:
    """LuxOS：Fan status changes 下「FANx: ... -> FAIL」类行。"""
    return len(re.findall(r"fan\d+\s*:\s*[^:\n]*->\s*fail", lower, re.I))


def _has_strong_hashboard_hw_evidence(lower: str) -> bool:
    """明确的芯片/链路硬件类报错（与单纯 Dead/无任务、历史 DNS 日志区分）"""
    return bool(
        re.search(r"no chips detected|nochipsdetected|\{err:i3\}", lower)
        or re.search(r"discovered 0 chips", lower)
        or re.search(r"tuner error.*hashboard|disabled hashboard", lower)
        or re.search(
            r"detected \d+\s+chips.*expected maximum|unexpected revision of chip|\{err:i2\}|\{err:i1\}",
            lower,
        )
        or ("nopic" in lower and "disabled hashboard" in lower)
        or re.search(r"\{err:e4\}|eeproms?\s+could not be decoded", lower, re.I)
        or re.search(r"doesn\x27t match chip count|number of responses \d+ of read_register", lower)
        or re.search(
            r"error_soc_init:\s*basic init failed|data load fail for chain|eeprom load ret\s*:\s*-1",
            lower,
            re.I,
        )
    )


def _learned_rule_skip_by_substrings(lower: str, rule: dict) -> bool:
    for skip_pat in rule.get("skip_if_log_matches") or []:
        if skip_pat and skip_pat.lower() in lower:
            return True
    return False


def _learned_rule_match_score(rule: dict, lower: str) -> bool:
    """
    学习规则匹配：
    - patterns_any：命中次数之和 >= min_total_hits（默认 1）
    - patterns_all：每个模式各自命中 >= min_each_hits（默认 1）
    二者可同时存在（更贴近「多条日志共同印证」的人工判读）。
    """
    m = rule.get("match") or {}
    pats_any = m.get("patterns_any") or []
    pats_all = m.get("patterns_all") or []
    if not pats_any and not pats_all:
        return False

    def _count_pat(p: str) -> int:
        try:
            return len(re.findall(p, lower, flags=re.I))
        except re.error:
            return lower.count(p.lower())

    if pats_any:
        total = sum(_count_pat(str(p)) for p in pats_any)
        min_total = int(m.get("min_total_hits", 1))
        if total < min_total:
            return False

    min_each = int(m.get("min_each_hits", 1))
    for p in pats_all:
        if _count_pat(str(p)) < min_each:
            return False
    return True


def _learned_rule_when(rule: dict) -> str:
    """
    unknown_only：仅在规则引擎主因仍为「未识别」时参与（走早期短路）。
    refine：在完整内置判定结束后执行，可在你指定条件下覆盖主因/方案（人工经验修正）。
    """
    w = (rule.get("when") or "").strip().lower()
    if w in ("refine", "refine_primary", "override"):
        return "refine"
    if w in ("unknown_only", "unknown"):
        return "unknown_only"
    if w == "":
        return "unknown_only" if rule.get("only_after_primary_unknown", True) else "refine"
    return "unknown_only"


def _evaluate_learned_extra_rules(
    lower: str,
    primary: str,
    curtail_sleep: bool,
) -> Optional[tuple]:
    """返回 (primary_cause, solutions_list) 或 None；仅 when=unknown_only 的规则。"""
    learned = _load_learned_json()
    sop = _merged_primary_sop()
    for rule in learned.get("extra_rules") or []:
        if not rule.get("enabled", False):
            continue
        if _learned_rule_when(rule) != "unknown_only":
            continue
        if rule.get("skip_if_curtailment_sleep") and curtail_sleep:
            continue
        if rule.get("only_after_primary_unknown") and primary != "未识别明确故障关键词":
            continue
        if _learned_rule_skip_by_substrings(lower, rule):
            continue
        if not _learned_rule_match_score(rule, lower):
            continue
        name = str(rule.get("primary_cause", "")).strip()
        sols = rule.get("solutions") or []
        if name and isinstance(sols, list):
            return (name, sols if sols else sop.get(name, ["请补充 fault_patterns_learned.json 中的 solutions"]))
    return None


def _safe_int_priority(rule: dict) -> int:
    try:
        return int(rule.get("priority", 0) or 0)
    except (TypeError, ValueError):
        return 0


def _try_learned_refine_primary(
    lower: str,
    primary: str,
    curtail_sleep: bool,
) -> Optional[Tuple[str, List[str], List[str]]]:
    """
    人工经验「修正」层：在固定规则已给出主因后，若命中 refine 规则则替换主因与方案。
    返回 (新主因, 新方案列表, 追加到 secondary 的提示语列表)。
    """
    learned = _load_learned_json()
    sop = _merged_primary_sop()
    refine_rules = [r for r in (learned.get("extra_rules") or []) if isinstance(r, dict)]
    refine_rules.sort(key=lambda r: (-_safe_int_priority(r), str(r.get("id", ""))))
    for rule in refine_rules:
        if not rule.get("enabled", False):
            continue
        if _learned_rule_when(rule) != "refine":
            continue
        if rule.get("skip_if_curtailment_sleep") and curtail_sleep:
            continue
        if _learned_rule_skip_by_substrings(lower, rule):
            continue

        only_if = [str(x) for x in (rule.get("only_if_primary_in") or []) if str(x).strip()]
        if only_if and not any(sub in primary for sub in only_if):
            continue

        not_if = [str(x) for x in (rule.get("not_if_primary_in") or []) if str(x).strip()]
        if not_if and any(sub in primary for sub in not_if):
            continue

        if not _learned_rule_match_score(rule, lower):
            continue

        name = str(rule.get("primary_cause", "")).strip()
        sols = rule.get("solutions") or []
        if not name or not isinstance(sols, list):
            continue
        sols_out = sols if sols else sop.get(name, ["请补充 fault_patterns_learned.json 中的 solutions"])
        hints = [str(x).strip() for x in (rule.get("secondary_hints") or []) if str(x).strip()]
        tag = str(rule.get("id") or "").strip()
        if tag:
            hints.insert(0, f"已由学习规则「{tag}」覆盖默认判定（人工经验）")
        return (name, list(sols_out), hints)
    return None


def load_diagnostic_few_shot_from_json() -> str:
    """从 fault_patterns_learned.json 读取少量人工范例，供 Ollama 叙事参考。"""
    learned = _load_learned_json()
    cases = learned.get("diagnostic_few_shot") or []
    if not isinstance(cases, list) or not cases:
        return ""
    parts: List[str] = []
    for i, c in enumerate(cases[:6], 1):
        if not isinstance(c, dict):
            continue
        title = str(c.get("title", f"范例{i}")).strip()
        excerpt = str(c.get("log_excerpt", "")).strip()[:1200]
        conclusion = str(c.get("your_conclusion", "")).strip()
        if not conclusion and not excerpt:
            continue
        parts.append(f"【{title}】\n你的结论：{conclusion}\n相关日志摘录：\n{excerpt}\n")
    return "\n".join(parts).strip()


def extract_ip(text: str) -> str:
    m = re.search(r"IP:\s*([0-9.]+)", text)
    return m.group(1) if m else ""


def _last_section_block(text: str, marker: str) -> str:
    """取最后一次出现的 [marker] 到下一个「时间戳] [节名]」之间的内容（含 JSON）。LuxOS 常见：行首为 [日期] [STATS]。"""
    pos = text.rfind(marker)
    if pos == -1:
        return ""
    rest = text[pos:]
    # 下一节多为：[2026-xx-xx hh:mm:ss] [STATS] / [DEVS] / …
    m = re.search(
        r"\n\[[0-9]{4}-[0-9]{2}-[0-9]{2}[^\]]*\]\s*\[(?:STATS|DEVS|POOLS|CONFIG|SUMMARY|MINER|LOG|INFO)",
        rest,
        re.I,
    )
    if m:
        return rest[: m.start()]
    return rest


def _devs_dead_alive_counts(text: str) -> Tuple[int, int]:
    """
    取末段 [DEVS] JSON 内各算力链 Status: Dead / Alive 出现次数（LuxOS 等）。
    仅解析 DEVS 节，避免 POOLS 的 Status: Alive 混入。
    """
    blk = _last_section_block(text, "[DEVS]")
    if not blk:
        return 0, 0
    b = blk.lower()
    dead = len(re.findall(r'"status"\s*:\s*"dead"', b))
    alive = len(re.findall(r'"status"\s*:\s*"alive"', b))
    return dead, alive


def _summary_ths_from_block(summary_blk: str) -> Optional[float]:
    """
    从单个 [SUMMARY] JSON 块取 TH/s：对 GHS 15m/30m/av/5s/1m、MHS 15m 等取最大值。
    避免旧逻辑「先匹配到 GHS 15m=0 就立刻返回」导致误判整批为 0 算力（15m 瞬时为 0 时 av/5s 仍可能有值）。
    """
    if not summary_blk:
        return None
    candidates: List[float] = []
    for pattern, divisor in [
        (r'"GHS\s+15m"\s*:\s*([0-9]+(?:\.[0-9]+)?)', 1000.0),
        (r'"GHS\s+30m"\s*:\s*([0-9]+(?:\.[0-9]+)?)', 1000.0),
        (r'"GHS\s+av"\s*:\s*([0-9]+(?:\.[0-9]+)?)', 1000.0),
        (r'"GHS\s+5s"\s*:\s*([0-9]+(?:\.[0-9]+)?)', 1000.0),
        (r'"GHS\s+1m"\s*:\s*([0-9]+(?:\.[0-9]+)?)', 1000.0),
        (r'"MHS\s+15m"\s*:\s*([0-9]+(?:\.[0-9]+)?)', 1_000_000.0),
    ]:
        m = re.search(pattern, summary_blk, re.I)
        if m:
            candidates.append(float(m.group(1)) / divisor)
    if not candidates:
        return None
    return max(candidates)


def extract_total_hashrate_ths(text: str) -> Optional[float]:
    """
    总算力（TH/s）。LUXminer 的 SUMMARY 用 GHS（数值÷1000 为 TH/s）；BOS 等用 MHS。
    禁止用「全文最后一个 MHS 15m」：DEVS 里每板一行，最后一板常为 0 会误判为 0 算力。
    """
    summary_blk = _last_section_block(text, "[SUMMARY]")
    st = _summary_ths_from_block(summary_blk) if summary_blk else None
    if st is not None and st > 0:
        return st
    # SUMMARY 全为 0 或无法从 SUMMARY 得到正值：再用各板 MHS 之和 / 全文 MHS 最大值兜底
    devs_blk = _last_section_block(text, "[DEVS]")
    if devs_blk:
        vals = re.findall(r'"MHS\s+15m"\s*:\s*([0-9]+(?:\.[0-9]+)?)', devs_blk, re.I)
        if vals:
            total_mhs = sum(float(x) for x in vals)
            if total_mhs > 0:
                return total_mhs / 1_000_000.0
    all_mhs = re.findall(r'"MHS\s+15m"\s*:\s*([0-9]+(?:\.[0-9]+)?)', text, re.I)
    if all_mhs:
        mx = max(float(x) for x in all_mhs)
        if mx > 0:
            return mx / 1_000_000.0
    # 曾从 SUMMARY 解析出数值（含全 0）→ 记为 0 算力；完全无 SUMMARY/无字段则跳过该文件
    if st is not None:
        return 0.0
    return None


def extract_nameplate_ths(text: str) -> Optional[float]:
    m = re.search(r'"NameplateTHS"\s*:\s*([0-9]+(?:\.[0-9]+)?)', text, re.I)
    if m:
        try:
            v = float(m.group(1))
            if v > 0:
                return v
        except ValueError:
            return None
    return None


def classify_hashrate_status(ths: float, nameplate_ths: Optional[float] = None) -> str:
    if ths <= ZERO_THS_EPSILON:
        return "无算力(0TH/s)"
    near_zero_ths = NEAR_ZERO_THS_ABS
    if nameplate_ths is not None and nameplate_ths > 0:
        near_zero_ths = max(near_zero_ths, nameplate_ths * NEAR_ZERO_NAMEPLATE_RATIO)
    if ths <= near_zero_ths:
        return f"无算力(近零<= {near_zero_ths:.2f}TH/s)"
    return f"低算力(<{THRESHOLD_THS:.0f}TH/s)"


def collect_evidence(text: str, limit: int = 12) -> str:
    lines = []
    priority_lines = []
    priority_patterns = [
        "received api command: curtail",
        "curtailment_mode",
        "skipping group connection due to curtailment",
        "curtailmode",
        "connection dropped due to pool inactivity",
        "stratum read failure",
        "connection closed by the server",
        "no chips detected",
        "nochipsdetected",
        "{err:i3}",
        "disabled hashboard",
        "discovered 0 chips",
        "discovered 26 chips",
        "expected 110",
        "zero mcr",
        "initialization of hashboard failed",
        "asic enumeration failure",
        "switching off the board",
        "soc init failed",
        "error_soc_init",
        "error_fan_lost",
        "fan lost",
        "reg crc error",
        "only find",
        "power off hashboard",
        "fail to read tsensor",
        "fail to read pic temp",
        "temperature sensor read failed",
        "ispowersupplyon",
        "fan status changes",
        "error_temp_too_high",
        "error_temp_too_low",
        "temp diff too high",
        "detected 2046",
        "expected maximum 110",
        "unexpected revision of chip",
        "nopic",
        '"status": "dead"',
        '"status": "alive"',
        "err:e4",
        "eeprom",
        "dummy backend",
        "doesn't match chip count",
        "hashchip",
        "temp diff too high",
        "wait for hashboard cool",
        "basic init failed",
        "data load fail for chain",
        "eeprom load ret",
    ]
    for line in text.splitlines():
        l = line.strip()
        if not l:
            continue
        lower = l.lower()
        if any(p in lower for p in priority_patterns):
            priority_lines.append(l)
            continue
        if "[error]" in lower or "error" in lower or "fail" in lower or "offline" in lower:
            lines.append(l)
            if len(lines) >= limit * 3:
                break
    merged = priority_lines[:limit] + lines[: max(0, limit - len(priority_lines[:limit]))]
    return "\n".join(merged[:limit])


def _truncate_line(s: str, max_len: int = 240) -> str:
    s = s.strip()
    if len(s) <= max_len:
        return s
    return s[: max_len - 1] + "…"


def _log_date_span_hint(text: str) -> str:
    """从日志中提取日期范围（YYYY-MM-DD），供叙事引用。"""
    dates = sorted(set(re.findall(r"\b(20[2-3]\d-\d{2}-\d{2})\b", text)))
    if not dates:
        return "未能从日志正文可靠提取日期范围（请以导出/采集时间为准）"
    if dates[0] == dates[-1]:
        return f"日志中出现的日期集中在 {dates[0]} 附近"
    return f"日志中出现的日期约从 {dates[0]} 至 {dates[-1]}"


def _extract_hb_numbers(text: str) -> List[int]:
    ids: List[int] = []
    for m in re.finditer(r"\{hb:(\d+)\}", text, flags=re.I):
        try:
            ids.append(int(m.group(1)))
        except ValueError:
            continue
    return sorted(set(ids))


def gather_categorized_log_lines(
    text: str,
    max_per_category: int = 10,
) -> Dict[str, List[str]]:
    """
    按主题从全量日志中抽取可引用原文行，用于「有理有据」叙事诊断。
    同一行可落入多个主题（板卡 + 总线等），便于交叉印证。
    """
    cats: Dict[str, List[str]] = {
        "board": [],
        "psu": [],
        "net": [],
        "temp_fan": [],
        "bus": [],
        "generic": [],
    }
    seen: set[str] = set()

    pat_board = re.compile(
        r"{err:i3}|no\s*chips\s*detected|nochipsdetected|disabled\s*hashboard|"
        r"discovered\s+\d+\s+chips|expected\s+110|only\s+find|tuner\s+error|"
        r"hashboard|expected\s+maximum\s+110|asic\s+enumeration|enumeration_reboot|"
        r"initialization\s+of\s+hashboard\s+failed|switching\s+off\s+the\s+board|"
        r'"status"\s*:\s*"dead"',
        re.I,
    )
    pat_psu = re.compile(
        r"psu|checksum|undervolt|over\s*voltage|overcurrent|"
        r"failed\s+to\s+(detect|set\s+up)\s+psu|dummy\s+backend",
        re.I,
    )
    pat_net = re.compile(
        r"failed\s+to\s+resolve|dns|stratum|no\s+stratum|pool\s+inactivity|"
        r"connection\s+closed|socket\s+(error|timeout|reset)|"
        r"stratum\s+read\s+failure|client\s+disconnected",
        re.I,
    )
    pat_temp_fan = re.compile(
        r"overtemp|temp\s+diff|error_temp|fail\s+to\s+read\s+tsensor|"
        r"pic\s+temp|fan\s+lost|error_fan|tachometer|temperature\s+sensor",
        re.I,
    )
    pat_bus = re.compile(
        r"probing\s+failed|no\s+such\s+device|i2c|uart|spi|reg\s+crc\s+error|"
        r"checksum.*i2c|communication",
        re.I,
    )

    def _push(cat: str, line: str) -> None:
        if len(cats[cat]) >= max_per_category:
            return
        key = f"{cat}|{line[:200]}"
        if key in seen:
            return
        seen.add(key)
        cats[cat].append(_truncate_line(line))

    for raw in text.splitlines():
        line = raw.strip()
        if not line or len(line) < 8:
            continue
        low = line.lower()
        matched = False
        if pat_board.search(low):
            _push("board", line)
            matched = True
        if pat_psu.search(low):
            _push("psu", line)
            matched = True
        if pat_net.search(low):
            _push("net", line)
            matched = True
        if pat_temp_fan.search(low):
            _push("temp_fan", line)
            matched = True
        if pat_bus.search(low):
            _push("bus", line)
            matched = True
        if not matched and (
            "[error]" in low
            or re.search(r"\berror\b", low)
            or "fail" in low
            or "warning" in low
        ):
            _push("generic", line)

    return cats


def _opening_summary_from_facts(
    primary: str,
    parsed_ths: Optional[float],
    nameplate_ths: Optional[float],
    hb_ids: List[int],
    board_lines: List[str],
) -> str:
    """基于规则主因 + 抽取事实生成开篇结论段（不编造日志未出现的编号）。"""
    ths_txt = f"{parsed_ths:.3f} TH/s" if parsed_ths is not None else "未知"
    np_txt = f"{nameplate_ths:.1f} TH/s" if nameplate_ths else None
    cap = f"识别总算力约 {ths_txt}"
    if np_txt:
        cap += f"，额定约 {np_txt}"

    hb_txt = ""
    if hb_ids:
        hb_txt = f"日志中明确出现算力板编号 HB:{','.join(str(x) for x in hb_ids)}。"

    i3_n = sum(1 for x in board_lines if re.search(r"{err:i3}|no\s*chips\s*detected|nochipsdetected", x, re.I))
    disc = [x for x in board_lines if re.search(r"discovered\s+\d+\s+chips", x, re.I)]

    tail = ""
    if i3_n >= 2:
        tail = "多条日志指向「检测不到芯片/ERR:I3」类致命错误，说明至少多块算力板已无法完成 ASIC 枚举。"
    elif i3_n == 1:
        tail = "日志出现「检测不到芯片/ERR:I3」类错误，至少一块算力板存在致命硬件或链路问题。"
    if disc:
        tail += " 另有日志显示某链「检出芯片数」远低于额定 110，属于严重残缺状态。"

    if not tail:
        tail = f"规则引擎将当前现象归类为「{primary}」；具体机理需结合下方原文摘录核对。"

    return f"综合规则判定与日志摘录：{cap}。{hb_txt}{tail}"


def _secondary_ranking_sentence(
    primary: str,
    secondary_joined: str,
    alt_joined: str,
    cats: Dict[str, List[str]],
) -> str:
    parts: List[str] = []
    if secondary_joined:
        parts.append(f"规则标注的伴生/次要因素包括：{secondary_joined}")
    if alt_joined:
        parts.append(f"候选需排除项：{alt_joined}")
    # 用是否有原文支撑来提示「伴生」
    if cats.get("psu") and "电源" not in primary and "PSU" not in primary.upper():
        parts.append("日志中可见电源通信/校验类原文，多为次生或连带现象，应在板卡与供电稳定后复测。")
    if cats.get("net") and "矿池" not in primary and "网络" not in primary:
        parts.append("日志中可见 DNS/Stratum/断连类原文，多为伴生问题；若总算力已为 0，应优先核对板卡硬件。")
    if not parts:
        return "未从规则侧列出额外次要标签；若日志仍有其他告警，请以原文摘录为准。"
    return " ".join(parts)


def build_narrative_report(
    text: str,
    rule_bundle: Dict[str, str],
    *,
    ip: str,
    model: str,
    nameplate_ths: Optional[float],
    parsed_ths: Optional[float],
    suggestions: List[str],
) -> str:
    """
    生成「有理有据」的结构化中文诊断叙述：结论 + 分主题原文摘录 + 主次说明 + 行动建议。
    仅使用日志中真实出现的行与规则引擎已有标签，不虚构具体日期/芯片数（摘录行内数字除外）。
    """
    primary = (rule_bundle.get("primary_cause") or "未识别明确故障关键词").strip()
    secondary = (rule_bundle.get("secondary_causes") or "").strip()
    alternate = (rule_bundle.get("alternate_causes") or "").strip()
    confidence = (rule_bundle.get("confidence") or "中").strip()

    cats = gather_categorized_log_lines(text, max_per_category=10)
    hb_ids = _extract_hb_numbers(text)
    date_hint = _log_date_span_hint(text)

    lines_out: List[str] = []
    lines_out.append(f"矿机（{ip}，型号 {model or '未知'}）")
    lines_out.append("")
    lines_out.append("【诊断依据概览】")
    lines_out.append(
        f"- 规则引擎主因：{primary}（置信度：{confidence}）"
    )
    lines_out.append(f"- 日志时间参考：{date_hint}")
    lines_out.append(_opening_summary_from_facts(primary, parsed_ths, nameplate_ths, hb_ids, cats["board"]))
    lines_out.append("")
    lines_out.append("【结论摘要（逻辑链）】")
    lines_out.append(
        _secondary_ranking_sentence(primary, secondary, alternate, cats)
    )
    lines_out.append("")

    def _section(title: str, key: str) -> None:
        rows = cats.get(key) or []
        if not rows:
            return
        lines_out.append(title)
        for i, row in enumerate(rows, 1):
            lines_out.append(f"  {i}. 日志摘录：{row}")
        lines_out.append("")

    _section("1. 算力板与芯片（与主因直接相关的原文）", "board")
    _section("2. 总线 / 传感器 / 通信类（可与板卡故障交叉印证）", "bus")
    _section("3. 电源与 PSU 通信（次生或诱因，需结合板卡现象）", "psu")
    _section("4. 网络与矿池 / Stratum（伴生或放大器）", "net")
    _section("5. 温度与风扇（保护链路与误报排查）", "temp_fan")
    _section("6. 其他告警行（泛化错误，供人工复核）", "generic")

    lines_out.append("【主次关系说明】")
    if primary in (
        "算力板全板失效（芯片未检出/初始化失败）",
        "算力板部分失效（芯片数量严重不足）",
        "单链算力板故障致 SOC 停机（芯片数异常/CRC/通信）",
        "多块算力板故障致 SOC 停机（芯片数严重不足/CRC）",
        "算力板严重异常（芯片识别/通信错误，NoPIC 可能无法单禁板）",
        "算力板芯片枚举不足（Hashchip 与 110 不符，已被禁用）",
    ):
        if cats.get("net"):
            lines_out.append(
                "当前主因归类为算力板/芯片侧。若同时存在矿池断连，多为「无有效算力可提交」或网络波动叠加；"
                "修复板卡与上电稳定后，应再观察 Stratum 是否恢复正常。"
            )
        else:
            lines_out.append("当前证据链主要围绕算力板/芯片与链路，请优先按板卡排障流程处理。")
    elif "矿池" in primary or "网络" in primary or "Stratum" in primary:
        lines_out.append(
            "当前主因归类为网络/矿池连接。若日志仍大量出现板级致命错误，则应将板卡问题提升为主矛盾，"
            "网络修复后仍需解决硬件侧。"
        )
    elif "限电" in primary or "休眠" in primary or "Curtail" in primary:
        lines_out.append("当前主因与策略休眠/限电相关；其他硬件告警可能是休眠前残留或唤醒失败连带，请结合时间线阅读摘录。")
    else:
        lines_out.append(
            "请以上方规则主因为主线，将其他主题的原文视为「支持/排除/伴生」证据；"
            "若摘录之间时间跨度较大，建议截取故障前后各 30 分钟日志复核。"
        )
    lines_out.append("")

    lines_out.append("【总结与行动建议】")
    for i, s in enumerate(suggestions[:12], 1):
        lines_out.append(f"  {i}. {s}")
    if not suggestions:
        lines_out.append("  （暂无结构化建议，请导出完整日志由人工复核。）")

    lines_out.append("")
    lines_out.append(
        "—— 说明：上文「日志摘录」均来自当前拉取的原文行；"
        "若需达到人工报告级别的时序分析（例如精确到某日某时的趋势），请保证导出日志包含完整时间戳并覆盖故障前后时段。"
    )
    return "\n".join(lines_out).strip()


def try_ollama_narrative_report(
    *,
    ip: str,
    model: str,
    primary: str,
    secondary: str,
    alternate: str,
    confidence: str,
    parsed_ths: Optional[float],
    nameplate_ths: Optional[float],
    evidence_text: str,
) -> Optional[str]:
    """
    可选：调用本地 Ollama 生成长文叙事诊断（更接近自然语言报告）。
    需设置环境变量 MINER_AI_NARRATIVE=1（或 true），且本机已启动 Ollama。
    """
    if os.environ.get("MINER_AI_NARRATIVE", "").strip() not in ("1", "true", "True"):
        return None
    ths_s = f"{parsed_ths:.6f}" if parsed_ths is not None else "null"
    np_s = f"{nameplate_ths:.3f}" if nameplate_ths else "null"
    few = load_diagnostic_few_shot_from_json()
    few_block = ""
    if few:
        few_block = (
            "\n\n你过去的人工诊断范例（请模仿其严谨性与主次结构，结论仍须以本次摘录为准）：\n"
            + few
            + "\n"
        )
    prompt = f"""你是资深矿机运维专家。请根据「规则引擎结论」与「日志原文摘录」写一份中文诊断报告。

硬性要求：
1) 不得编造日志里没有的 IP、型号、日期、芯片数量；所有数字与专有名词必须能在「日志原文摘录」或「规则引擎结论」中找到依据。
2) 结构包含：开篇结论段；分小节（算力板/电源/网络/其他）；「主次关系」一段；「行动建议」分条。
3) 每个小节至少引用 1 条原文摘录，格式为「摘录：」后跟短句。
4) 若某类证据缺失，明确写「当前摘录未覆盖」。

规则引擎结论：
- 主因：{primary}
- 置信度：{confidence}
- 次要：{secondary or "无"}
- 候选：{alternate or "无"}
矿机 IP：{ip}
型号：{model}
识别总算力(TH/s)：{ths_s}
额定 NameplateTHS（若有）：{np_s}

日志原文摘录（截断）：
{evidence_text[:12000]}
{few_block}
只输出 JSON：{{"narrative": "……"}} ，narrative 内用 \\n 换行。"""

    payload = {
        "model": OLLAMA_MODEL,
        "prompt": prompt.strip(),
        "stream": False,
        "format": "json",
        "options": {"temperature": 0.15},
    }
    req = urllib.request.Request(
        OLLAMA_URL,
        data=json.dumps(payload).encode("utf-8"),
        headers={"Content-Type": "application/json"},
        method="POST",
    )
    try:
        with urllib.request.urlopen(req, timeout=120) as resp:
            raw = resp.read().decode("utf-8", errors="ignore")
        data = json.loads(raw)
        text_resp = (data.get("response") or "").strip()
        if not text_resp:
            return None
        obj = json.loads(text_resp)
        nar = str(obj.get("narrative", "")).strip()
        return nar or None
    except (urllib.error.URLError, TimeoutError, json.JSONDecodeError, OSError, TypeError):
        return None


def rule_diagnose(text: str, parsed_hashrate_ths: Optional[float] = None) -> Dict[str, str]:
    lower = text.lower()
    primary = "未识别明确故障关键词"
    secondary: List[str] = []
    sop = _merged_primary_sop()
    priority_lane: Optional[str] = None
    from_rules_generic = False
    # CONFIG 断电：供传感器、风扇、Stratum 降权等多处复用
    power_supply_off_cfg = bool(re.search(r'"ispowersupplyon"\s*:\s*false', lower))

    # 最高优先级：限电/策略休眠
    # - 曾用「全文无 wakeup」判断 API curtail sleep，全天日志里只要有过一次唤醒整段即失效，已改为按行匹配含 sleep 的 curtail 指令。
    # - CONFIG 中单独的 "CurtailMode":"Sleep" 可能是固件默认，故仅在 IsPowerSupplyOn=false 或解析算力≤0 时与 CONFIG 组合采纳。
    curtail_sleep = bool(
        re.search(r"curtailment_mode\s*=\s*[\"']sleep[\"']", lower)
        or re.search(r"loading previous curtailment state:\s*['\"]sleep['\"]", lower)
        or re.search(r"received api command:\s*curtail[^\n]*\bsleep\b", lower)
        or "skipping group connection due to curtailment" in lower
        or "setting curtailment to sleep mode" in lower
        or "powering off the board due to curtailment" in lower
        or (
            re.search(r'"curtailmode"\s*:\s*"sleep"', lower)
            and (
                power_supply_off_cfg
                or (parsed_hashrate_ths is not None and parsed_hashrate_ths <= 0)
            )
        )
    )
    if curtail_sleep and parsed_hashrate_ths is not None and parsed_hashrate_ths > 0:
        curtail_sleep = False
    if curtail_sleep:
        primary = "限电/策略休眠（Curtailment Sleep）"
        secondary.append("退出休眠后若仍断连，再排查矿池/网络")
        priority_lane = "curtail"

    # 温度传感器 I2C/PIC 读取失败 → 误报超温或温差过大，触发 stop_mining（非风扇真实故障）
    # LuxOS 常见为 "Temperature Sensor Read Failed"，与 fail to read tsensor 措辞不同，需一并统计
    if primary == "未识别明确故障关键词":
        tsensor_fail = len(
            re.findall(
                r"fail to read tsensor|fail to read pic temp|fail to read.*pic.*temp|"
                r"temperature\s+sensor\s+read\s+failed|sensor\s+read\s+failed",
                lower,
            )
        )
        temp_protect = bool(
            re.search(
                r"error_temp_too_high|error_temp_too_low|temp diff too high|over max temp",
                lower,
            )
            or re.search(r"stop_mining:.*temp", lower)
            or re.search(r"\bnan\b", lower)
        )
        if (tsensor_fail >= 3 and temp_protect) or (tsensor_fail >= 10 and power_supply_off_cfg):
            primary = "温度传感器通信失败（误报超温/温差保护停机）"
            priority_lane = "tsensor"
            if "fan check passed" in lower:
                secondary.append("日志含 Fan check passed，可排除风扇本体故障")
            if power_supply_off_cfg:
                secondary.append(
                    "CONFIG 中 IsPowerSupplyOn 为 false：多为保护关机或未上电；请先查温度传感器排线/接口，再尝试唤醒或手动上电"
                )

    # 算力板间温差过大（temp diff too high）— 与「传感器读失败」区分：多为真实散热/单链异常，非总线 uart_trans 等版本串
    if primary == "未识别明确故障关键词":
        temp_diff_trip = bool(
            re.search(r"temp diff too high", lower, re.I)
            or re.search(r"stop_mining:\s*temp diff", lower, re.I)
            or (
                re.search(r"error_temp_too_low", lower, re.I)
                and "temp diff" in lower
            )
        )
        if temp_diff_trip:
            primary = "算力板间温差过大（温度保护停机）"
            priority_lane = "temp_diff"
            if "fan check passed" in lower:
                secondary.append("日志含 Fan check passed，可优先排查单链散热、风道堵塞与导热接触，而非风扇本体故障")
            secondary.append(
                "若出现 Wait for hashboard cool / Slept … diff =，属温差保护逻辑；重点处理高温链路与积灰"
            )

    # 原厂/部分固件：多链 EEPROM 加载失败 → ERROR_SOC_INIT basic init（非 ERR:E4 / 非 Braiins decode 句式）
    if primary == "未识别明确故障关键词":
        eeprom_chain_fail = bool(
            re.search(r"data load fail for chain", lower, re.I)
            or re.search(r"eeprom load ret\s*:\s*-1", lower, re.I)
            or re.search(r"fixture data load failed", lower, re.I)
            or re.search(r"eeprom error:\s*crc", lower, re.I)
            or re.search(r"got nothing", lower, re.I)
        )
        soc_basic_init = bool(
            re.search(r"error_soc_init:\s*basic init failed", lower, re.I)
            or re.search(r"stop_mining:\s*basic init failed", lower, re.I)
        )
        if soc_basic_init and (
            eeprom_chain_fail
            or re.search(r"load chain \d+ eeprom", lower, re.I)
        ):
            primary = "算力板 EEPROM 读取失败（多链初始化失败，SOC 无法启动）"
            priority_lane = "eeprom_init"
            secondary.append(
                "多链 EEPROM 无数据/校验/Fixture 失败会导致 SOC 无法完成基本初始化；优先排查排线与固件，再考虑控制板/算力板"
            )

    # 风扇丢失/数量不足 → stop_mining: fan lost / ERR:F3；后续 ERROR_SOC_INIT、power off hashboard 为保护连带，勿判为「单链 SOC」
    if primary == "未识别明确故障关键词":
        fan_err_f3 = bool(
            re.search(r"\{err:f3\}", lower)
            or re.search(r"less than required number of fans", lower)
            or re.search(r"detected\s+less\s+than\s+required\s+number\s+of\s+fans", lower)
        )
        fan_protect = bool(
            re.search(r"error_fan_lost|fan lost,\s*only find", lower)
            or re.search(r"stop_mining:\s*fan", lower)
            or fan_err_f3
            or (
                "fan lost" in lower
                and re.search(r"only find \d+\s*\(\s*<\s*\d+\s*\)", lower)
            )
        )
        if fan_protect:
            primary = "风扇异常（丢失/数量不足触发保护停机）"
            priority_lane = "fan"
            secondary.append("若同一时间段后出现 ERROR_SOC_INIT，多为风扇保护后的连带失败，主因仍为风扇检测")
            if fan_err_f3:
                secondary.append("日志含 ERR:F3 / less than required number of fans / miner will be paused，属于风扇数量不足触发的硬件保护停机")

    # LuxOS：Fan status changes 中 FANx 反复 -> FAIL（非 error_fan_lost 句式）+ CONFIG 断电 → 优先于矿池/Stratum 主因
    if primary == "未识别明确故障关键词":
        fan_lux_fail = _luxos_fan_fail_transitions(lower)
        if power_supply_off_cfg and fan_lux_fail >= LUXOS_FAN_FAIL_MIN_TRANSITIONS:
            primary = "风扇异常（LuxOS Fan status 反复 FAIL，伴电源关闭）"
            priority_lane = "fan_luxos"
            secondary.append(
                "CONFIG 中 IsPowerSupplyOn 为 false；矿池/Stratum 断连多为停挖或 pool inactivity，请先修复风扇与上电"
            )

    # Braiins OS / bosminer：算力板 EEPROM 无法解码（ERR:E4）阻止启动挖矿 — 优先于泛化「总线/通信」与矿池次要命中
    if primary == "未识别明确故障关键词":
        eeprom_decode_fail = bool(
            re.search(r"\{err:e4\}", lower, re.I)
            or re.search(r"err:e4\b", lower, re.I)
            or re.search(r"eeproms?\s+could not be decoded", lower, re.I)
            or re.search(r"cannot start miner.*eeprom", lower, re.I)
            or re.search(r"failed to decode\s+hashboard", lower, re.I)
            or re.search(r"eeprom\s+parser", lower, re.I)
        )
        if eeprom_decode_fail:
            primary = "算力板 EEPROM 解析失败（ERR:E4，无法启动挖矿）"
            priority_lane = "eeprom_e4"
            secondary.append(
                "多为指定算力板 EEPROM 数据损坏或硬件故障；优先重插该板排线/换槽位，仍失败则更换算力板"
            )
            if re.search(r"failed to (set up|detect) psu|failed to detect psu version", lower, re.I):
                secondary.append(
                    "日志同时有 PSU 协议检测失败：常在解决 EEPROM/板卡后复测；若仍存在再单独排查 PSU 与供电线"
                )

    # Braiins：Hashchip 响应数与 110 不符（Tuner 禁板）— 优先于「总线/通信」及 psu::i2c 中的裸 i2c 误命中
    if primary == "未识别明确故障关键词":
        hashchip_mismatch = bool(
            re.search(r"doesn\x27t match chip count", lower)
            or re.search(r"number of responses \d+ of read_register", lower)
        )
        if hashchip_mismatch and re.search(
            r"disabled hashboard|tuner error|init failed|start failed",
            lower,
            re.I,
        ):
            primary = "算力板芯片枚举不足（Hashchip 与 110 不符，已被禁用）"
            priority_lane = "hashchip"
            secondary.append(
                "多链芯片响应不足或无法通信，Tuner 已禁用对应算力板；优先重插排线/换槽位，仍失败则更换板卡"
            )
            if re.search(r"invalid checksum|psu::i2c|psu:.*checksum", lower, re.I):
                secondary.append(
                    "PSU I²C 校验失败可为次要因素，多在板卡与链路确认后再复测电源通信"
                )

    # Braiins OS / Tuner：ERR:I2 芯片数异常、ERR:I1 revision、NoPIC — 优先于大量历史 Failed to resolve 的「网络」计数
    if primary == "未识别明确故障关键词":
        tuner_hw = bool(
            re.search(
                r"detected \d+\s+chips.*expected maximum|unexpected revision of chip|\{err:i2\}|\{err:i1\}",
                lower,
            )
            and re.search(r"disabled hashboard|tuner error", lower)
        )
        nopic_hb = bool(
            "nopic" in lower
            and "disabled hashboard" in lower
            and re.search(r"can\x27t disable|can't disable", lower)
        )
        if tuner_hw or nopic_hb:
            primary = "算力板严重异常（芯片识别/通信错误，NoPIC 可能无法单禁板）"
            priority_lane = "tuner"
            secondary.append("若混有较早的 Failed to resolve，通常与当前 tuner 禁板无直接关系，以 HB 报错为主")

    # 原厂/部分固件：单链芯片数异常、CRC 等 → SOC init 失败、整机停挖（日志里 power off 多为关板，非 PSU）
    if primary == "未识别明确故障关键词":
        soc_sig = bool(
            re.search(r"error_soc_init|soc init failed", lower)
            or re.search(r"stop_mining.*soc init|failed to find the chip.*consecutive", lower)
        )
        # 勿单独用 power off hashboard：风扇保护停机也会关板，应与「Chain only find … asic」或 reg crc 等同现
        chain_sig = bool(
            re.search(r"chain\[?\d+\]?\s*:\s*find\s+\d+\s+asic", lower)
            or re.search(r"chain\s+\d+\s+only find \d+ asic", lower)
            or "reg crc error" in lower
        )
        if soc_sig and chain_sig:
            multi_only = len(
                re.findall(r"chain\s+\d+\s+only find\s+\d+\s+asic", lower)
            )
            if multi_only >= 2:
                primary = "多块算力板故障致 SOC 停机（芯片数严重不足/CRC）"
            else:
                primary = "单链算力板故障致 SOC 停机（芯片数异常/CRC/通信）"
            priority_lane = "soc"
            secondary.append("日志中「power off hashboard」为关断算力板动作，勿与 PSU 模块故障混淆")

    hb_ids = set(re.findall(r"\{hb:(\d+)\}", lower))
    no_chips_hits = len(re.findall(r"no chips detected|nochipsdetected|\{err:i3\}", lower))
    discovered_zero = len(re.findall(r"discovered 0 chips", lower))
    discovered_low = len(re.findall(r"discovered\s+([1-9]\d?)\s+chips.*expected 110", lower))

    # 算力板未检出/初始化失败（已判定为策略休眠时不再覆盖，避免把 Dead 误判为硬件坏板）
    if primary == "未识别明确故障关键词":
        if no_chips_hits > 0 and len(hb_ids) >= 2:
            primary = "算力板全板失效（芯片未检出/初始化失败）"
            secondary.append("疑似供电/排线/控制板链路问题")
            priority_lane = "hashboard"
        elif discovered_zero >= 2:
            primary = "算力板全板失效（芯片未检出/初始化失败）"
            priority_lane = "hashboard"
        elif no_chips_hits > 0 or discovered_low > 0:
            primary = "算力板部分失效（芯片数量严重不足）"
            priority_lane = "hashboard"

    # LuxOS [DEVS]：多块链 Dead、至少一块 Alive（无 tuner/no chips 等文本时仍可判定部分板失效）
    if primary == "未识别明确故障关键词":
        d_dead, d_alive = _devs_dead_alive_counts(text)
        if d_dead >= 1 and d_alive >= 1:
            primary = "算力板部分失效（多块链 Dead，仅部分链工作）"
            priority_lane = "hashboard"
            secondary.append(
                f"末段 [DEVS]：Dead={d_dead}，Alive={d_alive}；多为板卡/排线/槽位问题，建议交叉换槽位与检查供电接口"
            )

    # LuxOS：多块板初始化/枚举失败后关机，常仅剩一块板工作 → 总算力约为额定 1/3，非 0 算力
    init_fail = len(re.findall(r"initialization of hashboard failed", lower))
    enum_fail = len(re.findall(r"asic enumeration failure|enumeration_reboot", lower))
    if (
        primary == "未识别明确故障关键词"
        and init_fail >= 2
        and (enum_fail >= 2 or "switching off the board" in lower or "remain offline" in lower)
    ):
        primary = "部分算力板失效（初始化/ASIC枚举失败）"
        secondary.append("若 SUMMARY 仍有总算力，属低算力而非零算力")
        priority_lane = "init"

    # 单板运行但存在频繁 PLL 读错，通常为板级通信/稳定性问题
    if (
        primary == "未识别明确故障关键词"
        and len(re.findall(r"error reading chip pll", lower)) >= 3
    ):
        primary = "算力板部分失效（芯片数量严重不足）"
        secondary.append("日志存在频繁 PLL 读错，建议优先排查该板供电/排线/板卡老化")
        priority_lane = "pll"

    # 矿池/Stratum 频繁断连：CONFIG 断电且 Stratum 以 pool inactivity 为主时不列为主因（多为未挖矿/空闲踢线）
    stratum_hits = _stratum_unstable_total_hits(lower)
    pool_inactivity_hits = _stratum_pool_inactivity_hits(lower)
    stratum_dominated_by_pool_inactivity = (
        power_supply_off_cfg
        and stratum_hits >= STRATUM_UNSTABLE_MIN_HITS
        and pool_inactivity_hits >= 8
        and stratum_hits > 0
        and (pool_inactivity_hits / stratum_hits) >= STRATUM_POOL_INACTIVITY_DOMINANCE
    )
    if (
        primary == "未识别明确故障关键词"
        and not curtail_sleep
        and stratum_hits >= STRATUM_UNSTABLE_MIN_HITS
        and not _has_strong_hashboard_hw_evidence(lower)
        and not stratum_dominated_by_pool_inactivity
    ):
        primary = "矿池/Stratum连接不稳定（频繁断连）"
        secondary.append("算力板 Dead/低温常为未持续获得任务所致，先稳定矿池连接再判硬件")
        priority_lane = "stratum"

    if primary == "未识别明确故障关键词" and stratum_dominated_by_pool_inactivity:
        primary = "电源未开启或保护停机（CONFIG 断电）"
        priority_lane = "power_cfg"
        secondary.append(
            "日志中 pool inactivity / 断连多为未持续提交份额所致；请检查风扇、温度、外部策略与手动上电后再观察矿池连接"
        )

    # 从 fault_patterns_learned.json 加载的启用规则（自定义学习案例）
    if primary == "未识别明确故障关键词":
        learned_match = _evaluate_learned_extra_rules(lower, primary, curtail_sleep)
        if learned_match:
            primary, extra_sols = learned_match
            hits = _collect_rule_hits(lower)
            return _finalize_diagnosis_result(
                primary,
                secondary,
                "；".join(extra_sols),
                "learned",
                False,
                hits,
            )

    hardware_protect_shutdown = bool(
        re.search(r"\{err:f3\}", lower)
        or re.search(r"less than required number of fans", lower)
        or re.search(r"miner will be paused", lower)
        or re.search(r"halt reason:\s*hardware failure", lower)
    )

    hits = []
    for pattern, label, actions in RULES:
        c = len(re.findall(pattern, lower))
        # 已出现明确硬件保护停机时，弱化 "stratum.*disconnected" 的泛化网络噪音，
        # 保留 failed to resolve / dns error / socket timeout 等更直接网络证据。
        if label == "网络/矿池连接异常" and hardware_protect_shutdown:
            c -= len(re.findall(r"stratum.*disconnected", lower))
        if c:
            hits.append((c, label, actions))
    hits.sort(key=lambda x: x[0], reverse=True)

    if primary == "未识别明确故障关键词":
        if not hits:
            return _finalize_diagnosis_result(
                primary,
                [],
                "；".join(sop[primary]),
                priority_lane,
                False,
                hits,
            )
        primary = hits[0][1]
        from_rules_generic = True

    for _, label, _ in hits:
        if label != primary and label not in secondary:
            # 限电休眠时 fan 命中常为转速/日志噪音，不作为次因以免误导
            if primary == "限电/策略休眠（Curtailment Sleep）" and label == "风扇异常":
                continue
            # Stratum 不稳时 fan/泛化「板卡」常为误导：无明确芯片级报错则跳过
            if primary == "矿池/Stratum连接不稳定（频繁断连）":
                if label == "风扇异常":
                    continue
                if label == "算力板/芯片异常" and not _has_strong_hashboard_hw_evidence(lower):
                    continue
                if label == "电源异常" and not re.search(
                    r"failed to (set up|detect) psu|psu overcurrent|low voltage|undervolt", lower
                ):
                    continue
            if primary == "部分算力板失效（初始化/ASIC枚举失败）" and label in (
                "风扇异常",
                "电源异常",
                "算力板/芯片异常",
            ):
                continue
            if primary in (
                "单链算力板故障致 SOC 停机（芯片数异常/CRC/通信）",
                "多块算力板故障致 SOC 停机（芯片数严重不足/CRC）",
            ) and label in (
                "电源异常",
                "风扇异常",
                "总线/通信异常",
                "算力板/芯片异常",
            ):
                continue
            if primary == "温度传感器通信失败（误报超温/温差保护停机）" and label in (
                "风扇异常",
                "总线/通信异常",
                "算力板/芯片异常",
                "温度异常",
            ):
                continue
            if primary == "算力板严重异常（芯片识别/通信错误，NoPIC 可能无法单禁板）" and label in (
                "网络/矿池连接异常",
                "电源异常",
                "算力板/芯片异常",
            ):
                continue
            if primary == "风扇异常（丢失/数量不足触发保护停机）" and label in (
                "算力板/芯片异常",
                "总线/通信异常",
                "电源异常",
            ):
                continue
            if primary in (
                "风扇异常（LuxOS Fan status 反复 FAIL，伴电源关闭）",
                "电源未开启或保护停机（CONFIG 断电）",
            ) and label == "网络/矿池连接异常":
                continue
            if primary == "算力板 EEPROM 解析失败（ERR:E4，无法启动挖矿）" and label in (
                "总线/通信异常",
                "网络/矿池连接异常",
            ):
                continue
            if primary == "算力板芯片枚举不足（Hashchip 与 110 不符，已被禁用）" and label in (
                "总线/通信异常",
                "网络/矿池连接异常",
            ):
                continue
            if primary == "算力板间温差过大（温度保护停机）" and label == "总线/通信异常":
                continue
            if primary == "算力板 EEPROM 读取失败（多链初始化失败，SOC 无法启动）" and label in (
                "总线/通信异常",
                "矿机服务异常/接口超时",
            ):
                continue
            secondary.append(label)
        if len(secondary) >= 2:
            break

    refined = _try_learned_refine_primary(lower, primary, curtail_sleep)
    learned_refined = False
    if refined:
        new_primary, new_sols, sec_hints = refined
        primary = new_primary
        for h in sec_hints:
            if h and h not in secondary:
                secondary.append(h)
        solutions = "；".join(new_sols)
        from_rules_generic = False
        priority_lane = "learned"
        learned_refined = True

    if not learned_refined:
        if primary in sop:
            solutions = "；".join(sop[primary])
        else:
            actions_fallback: list[str] = []
            for _pat, lab, acts in RULES:
                if lab == primary and acts:
                    actions_fallback = list(acts)
                    break
            solutions = "；".join(actions_fallback or ["建议人工复核日志并做单机排障"])
    if hb_ids:
        solutions += f"；当前日志涉及 HB:{','.join(sorted(hb_ids))}"

    return _finalize_diagnosis_result(
        primary,
        secondary,
        solutions,
        priority_lane,
        from_rules_generic,
        hits,
    )


def _map_llm_confidence(raw: str) -> str:
    m = (raw or "").strip().lower()
    if m in ("high",):
        return "高"
    if m in ("medium", "med"):
        return "中"
    if m in ("low",):
        return "低"
    return "中"


def ollama_diagnose(ip: str, status: str, ths: float, evidence: str) -> Optional[Dict[str, str]]:
    prompt = f"""
你是矿机运维诊断助手。请基于下面日志证据分析“根因”和“可执行处理方案”。
要求：
1) 只输出 JSON，不要解释文字。
2) JSON 格式:
{{
  "cause": "一句话根因",
  "solutions": ["方案1", "方案2", "方案3"],
  "confidence": "high|medium|low"
}}
3) 方案必须具体可执行，不要空泛。

矿机IP: {ip}
状态: {status}
15分钟算力: {ths:.6f} TH/s
日志证据:
{evidence[:6000]}
""".strip()

    payload = {
        "model": OLLAMA_MODEL,
        "prompt": prompt,
        "stream": False,
        "format": "json",
        "options": {"temperature": 0.2},
    }
    req = urllib.request.Request(
        OLLAMA_URL,
        data=json.dumps(payload).encode("utf-8"),
        headers={"Content-Type": "application/json"},
        method="POST",
    )
    try:
        with urllib.request.urlopen(req, timeout=60) as resp:
            raw = resp.read().decode("utf-8", errors="ignore")
        data = json.loads(raw)
        text = (data.get("response") or "").strip()
        if not text:
            return None
        obj = json.loads(text)
        cause = str(obj.get("cause", "")).strip()
        solutions = obj.get("solutions", [])
        if isinstance(solutions, list):
            solutions_text = "；".join(str(x).strip() for x in solutions if str(x).strip())
        else:
            solutions_text = str(solutions).strip()
        if not cause:
            return None
        conf_cn = _map_llm_confidence(str(obj.get("confidence", "")).strip())
        return {"cause": cause, "solutions": solutions_text, "confidence": conf_cn}
    except (urllib.error.URLError, TimeoutError, json.JSONDecodeError, OSError):
        return None


def write_excel(rows: List[Dict[str, str]], output_path: Path) -> bool:
    if Workbook is None:
        return False
    wb = Workbook()
    headers = [
        "ip",
        "status",
        "ths_15m",
        "primary_cause",
        "confidence",
        "alternate_causes",
        "secondary_causes",
        "solutions",
        "evidence",
        "file",
    ]

    ws_all = wb.active
    ws_all.title = "全部异常"
    ws_all.append(headers)
    for r in rows:
        ws_all.append([r[h] for h in headers])

    ws_zero = wb.create_sheet("无算力")
    ws_zero.append(headers)
    for r in rows:
        if r["status"].startswith("无算力"):
            ws_zero.append([r[h] for h in headers])

    ws_low = wb.create_sheet("低算力")
    ws_low.append(headers)
    for r in rows:
        if not r["status"].startswith("无算力"):
            ws_low.append([r[h] for h in headers])

    for ws in [ws_all, ws_zero, ws_low]:
        ws.column_dimensions["A"].width = 16
        ws.column_dimensions["B"].width = 16
        ws.column_dimensions["C"].width = 14
        ws.column_dimensions["D"].width = 36
        ws.column_dimensions["E"].width = 8
        ws.column_dimensions["F"].width = 36
        ws.column_dimensions["G"].width = 28
        ws.column_dimensions["H"].width = 56
        ws.column_dimensions["I"].width = 72
        ws.column_dimensions["J"].width = 42

    try:
        wb.save(output_path)
    except PermissionError:
        alt = output_path.with_name(
            f"{output_path.stem}_{datetime.now().strftime('%Y%m%d_%H%M%S')}{output_path.suffix}"
        )
        wb.save(alt)
        print(f"Excel目标文件被占用，已另存为: {alt}")
    return True


def parse_args(argv: Optional[List[str]] = None) -> argparse.Namespace:
    p = argparse.ArgumentParser(
        description="扫描目录内矿机 *.txt 日志，筛低/零算力并生成规则+可选本地 LLM 诊断报表。",
    )
    p.add_argument(
        "log_dir",
        nargs="?",
        default=None,
        help="矿机日志所在目录（仅扫描该目录下 *.txt）。省略则使用脚本所在目录。",
    )
    p.add_argument(
        "-o",
        "--output-prefix",
        default="low_hashrate_ai_report",
        help="输出文件名前缀，生成 <前缀>.csv 与 <前缀>.xlsx（默认 low_hashrate_ai_report）",
    )
    return p.parse_args(argv)


def run_batch_diagnosis(
    log_root: Path,
    output_prefix: str = "low_hashrate_ai_report",
) -> Dict[str, Any]:
    """扫描 log_root 下 *.txt，生成低/零算力诊断报表（CSV + XLSX）。供 Web 与 CLI 共用。"""
    log_root = Path(log_root).expanduser().resolve()
    if not log_root.is_dir():
        raise FileNotFoundError(f"日志目录不存在: {log_root}")

    out_prefix = (output_prefix or "low_hashrate_ai_report").strip() or "low_hashrate_ai_report"
    rows: List[Dict[str, str]] = []
    llm_used = 0
    for file_path in sorted(log_root.glob("*.txt")):
        text = file_path.read_text(encoding="utf-8", errors="ignore")
        ip = extract_ip(text)
        ths = extract_total_hashrate_ths(text)
        nameplate_ths = extract_nameplate_ths(text)
        if ths is None:
            continue
        if ths >= THRESHOLD_THS:
            continue

        status = classify_hashrate_status(ths, nameplate_ths=nameplate_ths)
        evidence = collect_evidence(text, limit=12)

        rb = rule_diagnose(text, parsed_hashrate_ths=ths)
        primary_cause = rb["primary_cause"]
        secondary_causes = rb["secondary_causes"]
        solutions = rb["solutions"]
        confidence = rb.get("confidence", "中")
        alternate_causes = rb.get("alternate_causes", "")

        ai_result = None
        if (
            FORCE_LLM_OVERRIDE
            or primary_cause == "未识别明确故障关键词"
        ) and llm_used < MAX_AI_CALLS:
            ai_result = ollama_diagnose(ip, status, ths, evidence)
        if ai_result:
            llm_used += 1
            primary_cause = ai_result["cause"]
            secondary_causes = ""
            solutions = ai_result["solutions"] or "建议人工复核日志"
            confidence = ai_result.get("confidence", "中")
            alternate_causes = ""

        rows.append(
            {
                "ip": ip,
                "status": status,
                "ths_15m": f"{ths:.6f}",
                "primary_cause": primary_cause,
                "confidence": confidence,
                "alternate_causes": alternate_causes,
                "secondary_causes": secondary_causes,
                "solutions": solutions,
                "evidence": evidence,
                "file": file_path.name,
            }
        )

    output_csv = log_root / f"{out_prefix}.csv"
    with output_csv.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=[
                "ip",
                "status",
                "ths_15m",
                "primary_cause",
                "confidence",
                "alternate_causes",
                "secondary_causes",
                "solutions",
                "evidence",
                "file",
            ],
        )
        writer.writeheader()
        writer.writerows(rows)

    output_xlsx = log_root / f"{out_prefix}.xlsx"
    excel_ok = write_excel(rows, output_xlsx)

    zero_hard_count = sum(1 for r in rows if r["status"] == "无算力(0TH/s)")
    zero_near_count = sum(1 for r in rows if r["status"].startswith("无算力(近零"))
    zero_total_count = sum(1 for r in rows if r["status"].startswith("无算力"))
    low_count = sum(1 for r in rows if not r["status"].startswith("无算力"))

    return {
        "success": True,
        "log_dir": str(log_root),
        "csv_path": str(output_csv),
        "xlsx_path": str(output_xlsx) if excel_ok else None,
        "excel_ok": excel_ok,
        "row_count": len(rows),
        "zero_total_count": zero_total_count,
        "zero_hard_count": zero_hard_count,
        "zero_near_count": zero_near_count,
        "low_count": low_count,
        "llm_used": llm_used,
    }


def main(argv: Optional[List[str]] = None) -> None:
    args = parse_args(argv)
    log_root = Path(args.log_dir).expanduser().resolve() if args.log_dir else SCRIPT_DIR
    if not log_root.is_dir():
        print(f"错误：日志目录不存在: {log_root}", file=sys.stderr)
        sys.exit(1)

    try:
        r = run_batch_diagnosis(log_root, args.output_prefix)
    except Exception as e:
        print(f"错误：{e}", file=sys.stderr)
        sys.exit(1)

    print(f"异常矿机总数: {r['row_count']}")
    print(f"无算力总数: {r['zero_total_count']}")
    print(f"  其中0算力: {r['zero_hard_count']}")
    print(f"  其中近零算力: {r['zero_near_count']}")
    print(f"低算力(<50TH/s): {r['low_count']}")
    print(f"AI命中(使用本地LLM): {r['llm_used']}")
    print(f"日志目录: {r['log_dir']}")
    print(f"CSV: {r['csv_path']}")
    if r.get("excel_ok"):
        print(f"Excel: {r['xlsx_path']}")
    else:
        print("Excel: 未生成（请先安装 openpyxl）")


if __name__ == "__main__":
    main()
